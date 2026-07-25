import json
import os
import re

from openpyxl import load_workbook
from openpyxl import Workbook


class ExcelWriterAgent:

    def __init__(self, output_folder="output"):

        self.output_folder = output_folder

        os.makedirs(output_folder, exist_ok=True)

    def clean_json(self, response):

        response = response.strip()

        if response.startswith("```json"):
            response = response.replace("```json", "", 1)

        if response.startswith("```"):
            response = response.replace("```", "", 1)

        if response.endswith("```"):
            response = response[:-3]

        return response.strip()

    def clean_sheet_name(self, name):

        invalid = r'[\\/*?:\[\]]'

        name = re.sub(invalid, "_", name)

        return name[:31]

    def create_workbook(
            self,
            template_path,
            txt_file_name,
            ai_response
    ):

        tables = ai_response

        # Load uploaded template
        template = load_workbook(template_path)

        template_sheet = template.active

        headers = []

        for cell in template_sheet[1]:
            headers.append(cell.value)

        # New workbook
        workbook = Workbook()

        workbook.remove(workbook.active)

        for table in tables:

            table_name = table["table_name"]

            sheet = workbook.create_sheet(
                self.clean_sheet_name(table_name)
            )

            # Write headers
            for col, header in enumerate(headers, start=1):

                sheet.cell(
                    row=1,
                    column=col
                ).value = header

            # Write data
            row_no = 2

            for record in table["rows"]:

                for col, header in enumerate(headers, start=1):

                    sheet.cell(
                        row=row_no,
                        column=col
                    ).value = record.get(header, "")

                row_no += 1

        output_file = os.path.join(
            self.output_folder,
            txt_file_name.replace(".txt", ".xlsx")
        )

        workbook.save(output_file)

        return output_file